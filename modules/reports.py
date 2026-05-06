"""
ESG Reporting Module

This module handles generation of Excel and PDF reports with AI-powered insights.
"""

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime
import io
import os
from typing import Dict, Any, List, Tuple
import json

# Excel and PDF generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# AI Integration
try:
    import openai
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

from dotenv import load_dotenv

# Load environment variables
load_dotenv()


class ESGReporter:
    """Main class for generating ESG reports with AI insights."""
    
    def __init__(self):
        self.openai_client = None
        self.gemini_model = None
        self._initialize_ai_clients()
    
    def _initialize_ai_clients(self):
        """Initialize AI clients based on available API keys."""
        # OpenAI
        if OPENAI_AVAILABLE and os.getenv('OPENAI_API_KEY'):
            self.openai_client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        
        # Gemini
        if GEMINI_AVAILABLE and os.getenv('GEMINI_API_KEY'):
            genai.configure(api_key=os.getenv('GEMINI_API_KEY'))
            # Use the latest Gemini model
            try:
                self.gemini_model = genai.GenerativeModel('gemini-1.5-flash')
            except Exception as e:
                print(f"Failed to initialize gemini-1.5-flash, trying gemini-pro: {e}")
                try:
                    self.gemini_model = genai.GenerativeModel('gemini-pro')
                except Exception as e2:
                    print(f"Failed to initialize any Gemini model: {e2}")
                    self.gemini_model = None
    
    def generate_ai_insights(self, scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame]) -> str:
        """Generate AI-powered insights for the ESG report."""
        
        # Prepare data summary for AI
        data_summary = {
            "esg_scores": scores_dict,
            "data_summary": {
                "environmental_data_points": len(dfs.get('env', pd.DataFrame())),
                "social_data_points": len(dfs.get('social', pd.DataFrame())),
                "governance_data_points": len(dfs.get('gov', pd.DataFrame())),
            },
            "key_metrics": {
                "overall_esg_score": scores_dict.get('ESG', 0),
                "environmental_score": scores_dict.get('E', 0),
                "social_score": scores_dict.get('S', 0),
                "governance_score": scores_dict.get('G', 0),
            }
        }
        
        # Enhanced prompt for better Gemini responses
        prompt = f"""You are an expert ESG (Environmental, Social, Governance) sustainability consultant with 15+ years of experience. Analyze the following ESG data and provide comprehensive, actionable business insights.

**ESG PERFORMANCE DATA:**
- Overall ESG Score: {data_summary['key_metrics']['overall_esg_score']}/100
- Environmental Score: {data_summary['key_metrics']['environmental_score']}/100  
- Social Score: {data_summary['key_metrics']['social_score']}/100
- Governance Score: {data_summary['key_metrics']['governance_score']}/100

**DATA ANALYSIS SCOPE:**
- Environmental Data: {data_summary['data_summary']['environmental_data_points']} time periods analyzed
- Social Data: {data_summary['data_summary']['social_data_points']} time periods analyzed
- Governance Data: {data_summary['data_summary']['governance_data_points']} time periods analyzed

**REQUIRED ANALYSIS FRAMEWORK:**

1. **EXECUTIVE SUMMARY** (2-3 sentences)
   - Overall ESG performance rating and key takeaway

2. **STRENGTHS & ACHIEVEMENTS** (Top 3)
   - Identify strongest ESG categories with specific metrics
   - Highlight positive trends and achievements

3. **CRITICAL IMPROVEMENT AREAS** (Top 3)
   - Identify weakest ESG categories with specific concerns
   - Explain potential business risks and stakeholder impact

4. **STRATEGIC RECOMMENDATIONS** (3-4 specific actions)
   - Prioritized, actionable steps for ESG improvement
   - Include timelines and expected outcomes
   - Focus on quick wins and long-term sustainability

5. **RISK ASSESSMENT & OPPORTUNITIES**
   - Regulatory, reputational, and operational risks
   - Market opportunities and competitive advantages
   - Stakeholder engagement recommendations

**DELIVERABLE REQUIREMENTS:**
- Professional tone suitable for C-level executives
- Data-driven insights with specific score references
- Actionable recommendations with clear priorities
- Focus on business value and stakeholder impact
- Maximum 400 words for executive consumption

Provide your analysis now:"""
        
        # Try Gemini first (prioritized), then OpenAI
        try:
            if self.gemini_model:
                response = self.gemini_model.generate_content(prompt)
                if response.text:
                    return response.text.strip()
                else:
                    print("Gemini API returned empty response")
        
        except Exception as e:
            print(f"Gemini API error: {e}")
        
        try:
            if self.openai_client:
                response = self.openai_client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are an ESG sustainability expert providing actionable business insights."},
                        {"role": "user", "content": prompt}
                    ],
                    max_tokens=800,
                    temperature=0.7
                )
                return response.choices[0].message.content.strip()
        
        except Exception as e:
            print(f"OpenAI API error: {e}")
        
        # Fallback insights if AI is not available
        return self._generate_fallback_insights(scores_dict)
    
    def generate_detailed_esg_report(self, scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame], 
                                   company_name: str = "Company") -> str:
        """Generate a comprehensive ESG report using Gemini AI."""
        
        # Prepare detailed data for analysis
        detailed_data = self._prepare_detailed_data(scores_dict, dfs)
        
        prompt = f"""You are a senior ESG sustainability consultant preparing a comprehensive report for {company_name}. 

**COMPANY ESG PERFORMANCE OVERVIEW:**
- Overall ESG Score: {scores_dict.get('ESG', 0)}/100
- Environmental Score: {scores_dict.get('E', 0)}/100
- Social Score: {scores_dict.get('S', 0)}/100  
- Governance Score: {scores_dict.get('G', 0)}/100

**DETAILED METRICS ANALYSIS:**
{detailed_data}

**REPORT REQUIREMENTS:**
Generate a professional ESG sustainability report with the following sections:

1. **EXECUTIVE SUMMARY** (150 words)
   - Company ESG performance overview
   - Key achievements and challenges
   - Strategic recommendations summary

2. **ENVIRONMENTAL PERFORMANCE** (200 words)
   - Environmental score analysis and trends
   - Key environmental metrics and improvements
   - Sustainability initiatives and impact
   - Recommendations for environmental enhancement

3. **SOCIAL PERFORMANCE** (200 words)
   - Social score analysis and trends
   - Employee and community impact metrics
   - Social responsibility initiatives
   - Recommendations for social improvement

4. **GOVERNANCE PERFORMANCE** (200 words)
   - Governance score analysis and trends
   - Leadership and transparency metrics
   - Ethical business practices
   - Recommendations for governance enhancement

5. **STRATEGIC RECOMMENDATIONS** (150 words)
   - Prioritized action plan for ESG improvement
   - Timeline and resource requirements
   - Expected outcomes and business value
   - Stakeholder engagement strategy

6. **RISK ASSESSMENT & OPPORTUNITIES** (100 words)
   - ESG-related risks and mitigation strategies
   - Market opportunities and competitive advantages
   - Regulatory compliance considerations

**FORMATTING GUIDELINES:**
- Use professional business language
- Include specific data points and scores
- Provide actionable, prioritized recommendations
- Focus on business value and stakeholder impact
- Use bullet points and clear section headers
- Total word count: approximately 1,000 words

Generate the comprehensive ESG report now:"""

        try:
            if self.gemini_model:
                response = self.gemini_model.generate_content(prompt)
                if response.text:
                    return response.text.strip()
        except Exception as e:
            print(f"Gemini API error in detailed report: {e}")
        
        # Fallback to basic insights if Gemini fails
        return self.generate_ai_insights(scores_dict, dfs)
    
    def _prepare_detailed_data(self, scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame]) -> str:
        """Prepare detailed data summary for AI analysis."""
        data_summary = []
        
        # Environmental data details
        if 'env' in dfs and not dfs['env'].empty:
            env_df = dfs['env']
            data_summary.append("**ENVIRONMENTAL METRICS:**")
            for col in env_df.columns:
                if pd.api.types.is_numeric_dtype(env_df[col]):
                    current = env_df[col].iloc[-1] if len(env_df) > 0 else 0
                    avg = env_df[col].mean()
                    trend = "improving" if len(env_df) > 1 and current < env_df[col].iloc[0] else "stable/declining"
                    data_summary.append(f"- {col.replace('_', ' ').title()}: {current:.1f} (avg: {avg:.1f}, trend: {trend})")
        
        # Social data details
        if 'social' in dfs and not dfs['social'].empty:
            social_df = dfs['social']
            data_summary.append("\n**SOCIAL METRICS:**")
            for col in social_df.columns:
                if pd.api.types.is_numeric_dtype(social_df[col]):
                    current = social_df[col].iloc[-1] if len(social_df) > 0 else 0
                    avg = social_df[col].mean()
                    trend = "improving" if len(social_df) > 1 and current > social_df[col].iloc[0] else "stable/declining"
                    data_summary.append(f"- {col.replace('_', ' ').title()}: {current:.1f} (avg: {avg:.1f}, trend: {trend})")
        
        # Governance data details
        if 'gov' in dfs and not dfs['gov'].empty:
            gov_df = dfs['gov']
            data_summary.append("\n**GOVERNANCE METRICS:**")
            for col in gov_df.columns:
                if pd.api.types.is_numeric_dtype(gov_df[col]):
                    current = gov_df[col].iloc[-1] if len(gov_df) > 0 else 0
                    avg = gov_df[col].mean()
                    trend = "improving" if len(gov_df) > 1 and current > gov_df[col].iloc[0] else "stable/declining"
                    data_summary.append(f"- {col.replace('_', ' ').title()}: {current:.1f} (avg: {avg:.1f}, trend: {trend})")
        
        return "\n".join(data_summary) if data_summary else "Limited data available for analysis."
    
    def _generate_fallback_insights(self, scores_dict: Dict[str, Any]) -> str:
        """Generate fallback insights without AI."""
        esg_score = scores_dict.get('ESG', 0)
        e_score = scores_dict.get('E', 0)
        s_score = scores_dict.get('S', 0)
        g_score = scores_dict.get('G', 0)
        
        if esg_score >= 80:
            performance = "Excellent"
        elif esg_score >= 60:
            performance = "Good"
        elif esg_score >= 40:
            performance = "Fair"
        else:
            performance = "Needs Improvement"
        
        return f"""
        ESG Performance Assessment: {performance} ({esg_score}/100)
        
        Key Insights:
        • Environmental Performance: {e_score}/100
        • Social Performance: {s_score}/100  
        • Governance Performance: {g_score}/100
        
        Recommendations:
        • Focus on the lowest scoring category for maximum impact
        • Implement continuous monitoring and improvement processes
        • Engage stakeholders in sustainability initiatives
        • Consider industry best practices and benchmarking
        """
    
    def generate_excel_report(self, scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame], 
                            company_name: str = "Company") -> bytes:
        """Generate comprehensive Excel report with ESG data and charts."""
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16)
        
        # Summary Sheet
        summary_ws = wb.create_sheet("ESG Summary")
        self._create_summary_sheet(summary_ws, scores_dict, company_name, header_font, header_fill, title_font)
        
        # Environmental Sheet
        if 'env' in dfs and not dfs['env'].empty:
            env_ws = wb.create_sheet("Environmental Data")
            self._create_data_sheet(env_ws, dfs['env'], "Environmental KPIs", header_font, header_fill)
        
        # Social Sheet
        if 'social' in dfs and not dfs['social'].empty:
            social_ws = wb.create_sheet("Social Data")
            self._create_data_sheet(social_ws, dfs['social'], "Social KPIs", header_font, header_fill)
        
        # Governance Sheet
        if 'gov' in dfs and not dfs['gov'].empty:
            gov_ws = wb.create_sheet("Governance Data")
            self._create_data_sheet(gov_ws, dfs['gov'], "Governance KPIs", header_font, header_fill)
        
        # AI Insights Sheet
        insights_ws = wb.create_sheet("AI Insights")
        self._create_insights_sheet(insights_ws, scores_dict, dfs, company_name, header_font, header_fill)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def _create_summary_sheet(self, ws, scores_dict: Dict[str, Any], company_name: str, 
                            header_font, header_fill, title_font):
        """Create the summary sheet with ESG scores."""
        
        # Title
        ws['A1'] = f"ESG Sustainability Report - {company_name}"
        ws['A1'].font = title_font
        
        # Date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # ESG Scores Section
        ws['A4'] = "ESG Performance Summary"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        
        # Headers
        headers = ['Category', 'Score', 'Grade', 'Weight (%)', 'Contribution']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        categories = [
            ('Environmental', scores_dict.get('E', 0), 'E'),
            ('Social', scores_dict.get('S', 0), 'S'),
            ('Governance', scores_dict.get('G', 0), 'G'),
            ('Overall ESG', scores_dict.get('ESG', 0), 'ESG')
        ]
        
        for row, (category, score, key) in enumerate(categories, 6):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=f"{score}/100")
            
            # Grade calculation
            if score >= 95:
                grade = "A+"
            elif score >= 90:
                grade = "A"
            elif score >= 85:
                grade = "A-"
            elif score >= 80:
                grade = "B+"
            elif score >= 75:
                grade = "B"
            elif score >= 70:
                grade = "B-"
            elif score >= 65:
                grade = "C+"
            elif score >= 60:
                grade = "C"
            elif score >= 55:
                grade = "C-"
            elif score >= 50:
                grade = "D"
            else:
                grade = "F"
            
            ws.cell(row=row, column=3, value=grade)
            
            if key in scores_dict.get('weights', {}):
                ws.cell(row=row, column=4, value=f"{scores_dict['weights'][key]}%")
                contribution = scores_dict.get('contributions', {}).get(f'{key}_contribution', 0)
                ws.cell(row=row, column=5, value=f"{contribution:.2f}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_data_sheet(self, ws, df: pd.DataFrame, title: str, header_font, header_fill):
        """Create a data sheet with KPI information."""
        
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        
        # Add data starting from row 3
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style headers
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_insights_sheet(self, ws, scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame], 
                             company_name: str, header_font, header_fill):
        """Create AI insights sheet."""
        
        # Title
        ws['A1'] = f"AI-Generated Insights - {company_name}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Generate insights
        insights = self.generate_ai_insights(scores_dict, dfs)
        
        # Add insights to sheet
        ws['A3'] = "Executive Summary"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        
        # Split insights into paragraphs and add to sheet
        paragraphs = insights.split('\n\n')
        row = 4
        for paragraph in paragraphs:
            if paragraph.strip():
                ws.cell(row=row, column=1, value=paragraph.strip())
                ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
                row += 2
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 100
    
    def generate_pdf_report(self, scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame], 
                          company_name: str = "Company") -> bytes:
        """Generate comprehensive PDF report with ESG data and insights."""
        
        # Create PDF buffer
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=72, leftMargin=72, 
                              topMargin=72, bottomMargin=18)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
        
        # Build story
        story = []
        
        # Title
        title = Paragraph(f"ESG Sustainability Report<br/>{company_name}", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Date
        date_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(date_text, normal_style))
        story.append(Spacer(1, 20))
        
        # ESG Summary Table
        story.append(Paragraph("ESG Performance Summary", heading_style))
        
        # Create summary table
        summary_data = [
            ['Category', 'Score', 'Grade', 'Weight (%)'],
            ['Environmental', f"{scores_dict.get('E', 0)}/100", self._get_grade(scores_dict.get('E', 0)), 
             f"{scores_dict.get('weights', {}).get('E', 0)}%"],
            ['Social', f"{scores_dict.get('S', 0)}/100", self._get_grade(scores_dict.get('S', 0)), 
             f"{scores_dict.get('weights', {}).get('S', 0)}%"],
            ['Governance', f"{scores_dict.get('G', 0)}/100", self._get_grade(scores_dict.get('G', 0)), 
             f"{scores_dict.get('weights', {}).get('G', 0)}%"],
            ['Overall ESG', f"{scores_dict.get('ESG', 0)}/100", self._get_grade(scores_dict.get('ESG', 0)), '100%']
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # AI Insights
        story.append(Paragraph("AI-Generated Insights", heading_style))
        insights = self.generate_ai_insights(scores_dict, dfs)
        story.append(Paragraph(insights, normal_style))
        story.append(Spacer(1, 20))
        
        # Key Metrics Summary
        story.append(Paragraph("Key Metrics Summary", heading_style))
        
        # Environmental metrics
        if 'env' in dfs and not dfs['env'].empty:
            env_df = dfs['env']
            env_summary = []
            for col in env_df.columns:
                if not env_df[col].empty:
                    avg_val = env_df[col].mean()
                    env_summary.append(f"• {col.replace('_', ' ').title()}: {avg_val:.2f}")
            
            if env_summary:
                story.append(Paragraph("Environmental Metrics:", normal_style))
                for metric in env_summary:
                    story.append(Paragraph(metric, normal_style))
                story.append(Spacer(1, 10))
        
        # Social metrics
        if 'social' in dfs and not dfs['social'].empty:
            social_df = dfs['social']
            social_summary = []
            for col in social_df.columns:
                if not social_df[col].empty:
                    avg_val = social_df[col].mean()
                    social_summary.append(f"• {col.replace('_', ' ').title()}: {avg_val:.2f}")
            
            if social_summary:
                story.append(Paragraph("Social Metrics:", normal_style))
                for metric in social_summary:
                    story.append(Paragraph(metric, normal_style))
                story.append(Spacer(1, 10))
        
        # Governance metrics
        if 'gov' in dfs and not dfs['gov'].empty:
            gov_df = dfs['gov']
            gov_summary = []
            for col in gov_df.columns:
                if not gov_df[col].empty:
                    avg_val = gov_df[col].mean()
                    gov_summary.append(f"• {col.replace('_', ' ').title()}: {avg_val:.2f}")
            
            if gov_summary:
                story.append(Paragraph("Governance Metrics:", normal_style))
                for metric in gov_summary:
                    story.append(Paragraph(metric, normal_style))
        
        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return pdf_buffer.getvalue()
    
    def _get_grade(self, score: float) -> str:
        """Convert score to letter grade."""
        if score >= 95:
            return "A+"
        elif score >= 90:
            return "A"
        elif score >= 85:
            return "A-"
        elif score >= 80:
            return "B+"
        elif score >= 75:
            return "B"
        elif score >= 70:
            return "B-"
        elif score >= 65:
            return "C+"
        elif score >= 60:
            return "C"
        elif score >= 55:
            return "C-"
        elif score >= 50:
            return "D"
        else:
            return "F"


# Convenience functions for direct use
def generate_excel_report(scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame], 
                         company_name: str = "Company") -> bytes:
    """Generate Excel report with ESG data and AI insights."""
    reporter = ESGReporter()
    return reporter.generate_excel_report(scores_dict, dfs, company_name)


def generate_pdf_report(scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame], 
                       company_name: str = "Company") -> bytes:
    """Generate PDF report with ESG data and AI insights."""
    reporter = ESGReporter()
    return reporter.generate_pdf_report(scores_dict, dfs, company_name)


def generate_ai_insights(scores_dict: Dict[str, Any], dfs: Dict[str, pd.DataFrame]) -> str:
    """Generate AI-powered insights for ESG data."""
    reporter = ESGReporter()
    return reporter.generate_ai_insights(scores_dict, dfs)
